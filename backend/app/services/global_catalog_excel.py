from __future__ import annotations

from io import BytesIO
from typing import Any

from fastapi import HTTPException
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

COLUMNS = ("canonical_name", "brand", "category", "barcode_sku", "package_size", "package_type", "aliases", "active", "image_filename")
MAX_ROWS = 2000


def build_template() -> bytes:
    book = Workbook()
    sheet = book.active
    sheet.title = "Ürünler"
    sheet.append(COLUMNS)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="186A5A")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"
    widths = (30, 22, 22, 18, 16, 16, 42, 12, 28)
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    example = book.create_sheet("Örnek")
    example.append(COLUMNS)
    example.append(("Coca Cola 1L", "Coca Cola", "Gazlı İçecek", "5449000000996", "1", "L", "Coca Cola 1 Lt;Coca-Cola 1L", "true", "coca-cola-1l.png"))
    notes = book.create_sheet("Açıklamalar")
    notes.append(("Sütun", "Açıklama", "Örnek"))
    for cell in notes[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="186A5A")
    descriptions = {
        "canonical_name": "Zorunlu. Ürünün kanonik adı.", "brand": "Mevcut global marka adı.", "category": "Mevcut global kategori adı.",
        "barcode_sku": "Barkod/SKU. Varsa benzersiz olmalıdır.", "package_size": "Paket miktarı.", "package_type": "Paket birimi/türü.",
        "aliases": "Noktalı virgül (;) ile ayrılmış alternatif adlar.", "active": "true/false, evet/hayır, 1/0.", "image_filename": "İsteğe bağlı: Toplu Görsel Yükle paketindeki dosya adı.",
    }
    for column in COLUMNS:
        notes.append((column, descriptions[column], "Coca Cola 1L" if column == "canonical_name" else ""))
    notes.freeze_panes = "A2"
    notes.column_dimensions["A"].width = 22; notes.column_dimensions["B"].width = 75; notes.column_dimensions["C"].width = 30
    stream = BytesIO(); book.save(stream); return stream.getvalue()


def _text(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def _boolean(value: Any) -> bool | None:
    text = _text(value).casefold()
    if text in {"", "true", "1", "evet", "yes"}: return True
    if text in {"false", "0", "hayır", "hayir", "no"}: return False
    return None


def parse_workbook(content: bytes) -> list[dict[str, Any]]:
    try:
        book = load_workbook(BytesIO(content), read_only=True, data_only=False)
    except Exception as exc:
        raise HTTPException(422, "Geçersiz veya bozuk Excel dosyası.") from exc
    if "Ürünler" not in book.sheetnames:
        raise HTTPException(422, "Çalışma kitabında 'Ürünler' sayfası bulunamadı.")
    sheet = book["Ürünler"]
    header = [_text(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    if tuple(header[:len(COLUMNS)]) != COLUMNS:
        raise HTTPException(422, "Excel başlıkları şablonla eşleşmiyor.")
    rows = []
    for number, values in enumerate(sheet.iter_rows(min_row=2, values_only=False), 2):
        if number > MAX_ROWS + 1: raise HTTPException(413, f"En fazla {MAX_ROWS} ürün satırı yüklenebilir.")
        if all(cell.value in (None, "") for cell in values[:len(COLUMNS)]): continue
        if any(cell.data_type == "f" for cell in values[:len(COLUMNS)]):
            rows.append({"row": number, "status": "invalid", "error": "Formül içeren hücreler desteklenmez."}); continue
        record = {column: _text(values[index].value) for index, column in enumerate(COLUMNS)}
        record["row"] = number
        active = _boolean(record["active"])
        if not record["canonical_name"]: record.update(status="invalid", error="canonical_name zorunludur.")
        elif active is None: record.update(status="invalid", error="active true/false olmalıdır.")
        else: record["active"] = active
        rows.append(record)
    if not rows: raise HTTPException(422, "Ürünler sayfasında aktarılacak satır bulunamadı.")
    return rows
