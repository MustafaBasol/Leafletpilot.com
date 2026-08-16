"""Excel parsing for the Platform Admin market catalog import pilot.

Mirrors the openpyxl mechanics of ``global_catalog_excel.py`` (bounded rows,
formula rejection, boolean parsing) but targets one market's price/stock
sheet instead of the canonical global catalog. Header matching is
alias-based rather than fixed-position so a customer's spreadsheet (English
or Turkish headers, reordered columns, extra notes columns) still parses.
"""
from __future__ import annotations

from io import BytesIO
from typing import Any

from fastapi import HTTPException
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.services.product_identity import normalize_words

COLUMNS = (
    "product_name",
    "brand",
    "barcode",
    "package_amount",
    "package_unit",
    "package_type",
    "package_size",
    "category",
    "price",
    "promo_price",
    "currency",
    "sku",
    "image_url",
)
REQUIRED_COLUMNS = ("product_name",)
MAX_ROWS = 3000
SHEET_NAME = "Ürünler"

# Deterministic alias mapping. Keys are normalized (Turkish-folded, casefold,
# collapsed-space) header text; values are the canonical column key above.
# Keys are run through the same Turkish-folding normalizer as the header
# cells (normalize_words), which also collapses underscores to spaces — so
# every alias key below is written in its space-separated form.
_RAW_HEADER_ALIASES: dict[str, str] = {
    "product_name": "product_name",
    "urun adi": "product_name",
    "urun": "product_name",
    "urun ismi": "product_name",
    "name": "product_name",
    "brand": "brand",
    "marka": "brand",
    "barcode": "barcode",
    "barkod": "barcode",
    "barkod sku": "barcode",
    "ean": "barcode",
    "package_amount": "package_amount",
    "miktar": "package_amount",
    "amount": "package_amount",
    "package_unit": "package_unit",
    "birim": "package_unit",
    "unit": "package_unit",
    "package_type": "package_type",
    "paket tipi": "package_type",
    "ambalaj": "package_type",
    "package_size": "package_size",
    "gramaj": "package_size",
    "paket boyutu": "package_size",
    "size": "package_size",
    "category": "category",
    "kategori": "category",
    "price": "price",
    "fiyat": "price",
    "normal fiyat": "price",
    "regular_price": "price",
    "promo_price": "promo_price",
    "indirimli fiyat": "promo_price",
    "kampanya fiyati": "promo_price",
    "promo fiyat": "promo_price",
    "currency": "currency",
    "para birimi": "currency",
    "doviz": "currency",
    "sku": "sku",
    "stok kodu": "sku",
    "internal code": "sku",
    "image_url": "image_url",
    "gorsel url": "image_url",
    "gorsel": "image_url",
    "image": "image_url",
}


def _header_key(value: Any) -> str:
    return normalize_words(str(value)) if value is not None else ""


# Normalize alias keys through the same folding used for header cells so
# "product_name" and "product name" (and their Turkish-cased variants) match.
_HEADER_ALIASES: dict[str, str] = {
    normalize_words(key): value for key, value in _RAW_HEADER_ALIASES.items()
}


def build_template() -> bytes:
    book = Workbook()
    sheet = book.active
    sheet.title = SHEET_NAME
    sheet.append(COLUMNS)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="186A5A")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"
    widths = (30, 20, 18, 14, 12, 14, 14, 20, 10, 12, 10, 16, 32)
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    example = book.create_sheet("Örnek")
    example.append(COLUMNS)
    example_rows = (
        ("Ülker Çokomel", "Ülker", "8690504039025", "24", "g", "packet", "", "Atıştırmalık", "12.90", "", "EUR", "", ""),
        ("Coca Cola 1,5 LT", "Coca Cola", "", "1.5", "l", "bottle", "", "Gazlı İçecek", "2.50", "2.10", "EUR", "", ""),
        ("Sütaş Ayran 1L", "Sütaş", "", "1", "l", "bottle", "", "Süt Ürünleri", "1.80", "", "EUR", "", ""),
        ("Eti Burçak 131 g", "Eti", "", "", "", "", "131gr", "Atıştırmalık", "3.20", "", "EUR", "", ""),
    )
    for row in example_rows:
        example.append(row)
    notes = book.create_sheet("Açıklamalar")
    notes.append(("Sütun", "Açıklama", "Zorunlu mu?"))
    for cell in notes[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="186A5A")
    descriptions = {
        "product_name": ("Ürünün adı. (Ürün Adı da kabul edilir)", "Evet"),
        "brand": ("Marka adı. (Marka)", "Hayır"),
        "barcode": ("Barkod veya SKU. (Barkod)", "Hayır"),
        "package_amount": ("Sayısal paket miktarı, örn. 100.", "Hayır"),
        "package_unit": ("g, kg, ml, cl, l, pcs. (Birim)", "Hayır"),
        "package_type": ("bottle, can, box, bag, jar, packet.", "Hayır"),
        "package_size": ("Serbest metin paket boyutu, örn. 100gr, 1lt, 33cl. (Gramaj)", "Hayır"),
        "category": ("Mevcut kategori adı. (Kategori)", "Hayır"),
        "price": ("Normal satış fiyatı. (Fiyat)", "Hayır"),
        "promo_price": ("Kampanya/indirimli fiyat. (İndirimli Fiyat)", "Hayır"),
        "currency": ("EUR, TRY, USD, GBP, CHF. Boşsa marketin para birimi kullanılır.", "Hayır"),
        "sku": ("İç stok kodu.", "Hayır"),
        "image_url": ("Herkese açık http(s) görsel adresi.", "Hayır"),
    }
    for column in COLUMNS:
        description, required = descriptions[column]
        notes.append((column, description, required))
    notes.freeze_panes = "A2"
    notes.column_dimensions["A"].width = 18
    notes.column_dimensions["B"].width = 60
    notes.column_dimensions["C"].width = 14
    stream = BytesIO()
    book.save(stream)
    return stream.getvalue()


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _resolve_headers(header_cells: list[Any]) -> dict[int, str]:
    """Map spreadsheet column index -> canonical column key via aliases."""
    resolved: dict[int, str] = {}
    for index, raw in enumerate(header_cells):
        key = _HEADER_ALIASES.get(_header_key(raw))
        if key and key not in resolved.values():
            resolved[index] = key
    return resolved


def parse_workbook(content: bytes) -> list[dict[str, Any]]:
    try:
        book = load_workbook(BytesIO(content), read_only=True, data_only=False)
    except Exception as exc:
        raise HTTPException(422, "Geçersiz veya bozuk Excel dosyası.") from exc
    sheet = book[SHEET_NAME] if SHEET_NAME in book.sheetnames else book[book.sheetnames[0]]
    rows_iter = sheet.iter_rows(min_row=1, max_row=1)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise HTTPException(422, "Çalışma kitabında başlık satırı bulunamadı.") from None
    header_values = [cell.value for cell in header_row]
    column_map = _resolve_headers(header_values)
    if "product_name" not in column_map.values():
        raise HTTPException(
            422,
            "Excel başlıklarında zorunlu 'product_name' (Ürün Adı) sütunu bulunamadı.",
        )
    rows: list[dict[str, Any]] = []
    for number, values in enumerate(sheet.iter_rows(min_row=2, values_only=False), 2):
        if number > MAX_ROWS + 1:
            raise HTTPException(413, f"En fazla {MAX_ROWS} ürün satırı yüklenebilir.")
        relevant_cells = [values[index] for index in column_map if index < len(values)]
        if all(cell.value in (None, "") for cell in relevant_cells):
            continue
        if any(cell.data_type == "f" for cell in relevant_cells):
            rows.append({"row": number, "status": "invalid", "error": "Formül içeren hücreler desteklenmez."})
            continue
        record: dict[str, Any] = {column: "" for column in COLUMNS}
        for index, column in column_map.items():
            if index < len(values):
                record[column] = _text(values[index].value)
        record["row"] = number
        rows.append(record)
    if not rows:
        raise HTTPException(422, "Sayfada aktarılacak satır bulunamadı.")
    return rows
