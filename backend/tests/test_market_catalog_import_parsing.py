"""Pure parsing/normalization tests for the Phase 28D market catalog import.

No database required.
"""
from io import BytesIO

import pytest
from openpyxl import Workbook

from app.services.market_catalog_excel import COLUMNS, build_template, parse_workbook
from app.services.market_catalog_import import _normalize_row, _parse_decimal


def _workbook(headers, rows):
    book = Workbook()
    sheet = book.active
    sheet.title = "Ürünler"
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    stream = BytesIO()
    book.save(stream)
    return stream.getvalue()


def test_template_headers_are_recognized_by_the_parser():
    from openpyxl import load_workbook

    content = build_template()
    book = load_workbook(BytesIO(content))
    assert "Ürünler" in book.sheetnames
    assert "Örnek" in book.sheetnames
    # The blank "Ürünler" sheet has no data rows to parse; the "Örnek" sheet
    # demonstrates the same headers with filled-in example rows.
    example_content = _workbook(list(COLUMNS), list(book["Örnek"].iter_rows(min_row=2, values_only=True)))
    rows = parse_workbook(example_content)
    assert rows
    assert rows[0]["product_name"]


def test_english_headers_parse_by_position():
    content = _workbook(list(COLUMNS), [("Coca Cola 1L", "Coca Cola", "5449000000996", "1", "l", "bottle", "", "İçecek", "2.50", "2.10", "EUR", "SKU1", "https://example.test/coca.png")])
    rows = parse_workbook(content)
    assert len(rows) == 1
    assert rows[0]["product_name"] == "Coca Cola 1L"
    assert rows[0]["barcode"] == "5449000000996"
    assert rows[0]["price"] == "2.50"


def test_turkish_aliases_and_reordered_columns_parse():
    headers = ["Ürün Adı", "Marka", "Barkod", "Fiyat", "İndirimli Fiyat", "Gramaj"]
    content = _workbook(headers, [("Ülker Çokomel", "Ülker", "8690504039025", "12,90", "", "24 gr")])
    rows = parse_workbook(content)
    assert len(rows) == 1
    row = rows[0]
    assert row["product_name"] == "Ülker Çokomel"
    assert row["brand"] == "Ülker"
    assert row["barcode"] == "8690504039025"
    assert row["price"] == "12,90"
    assert row["package_size"] == "24 gr"


def test_missing_product_name_column_is_a_structural_error():
    content = _workbook(["Marka", "Fiyat"], [("Ülker", "1")])
    with pytest.raises(Exception) as excinfo:
        parse_workbook(content)
    assert "product_name" in str(excinfo.value)


def test_formula_cells_are_rejected_per_row():
    book = Workbook()
    sheet = book.active
    sheet.title = "Ürünler"
    sheet.append(["product_name", "price"])
    sheet.append(["Coca Cola 1L", "=1+1"])
    stream = BytesIO()
    book.save(stream)
    rows = parse_workbook(stream.getvalue())
    assert rows[0]["status"] == "invalid"
    assert "Formül" in rows[0]["error"]


def test_blank_rows_are_skipped():
    content = _workbook(["product_name", "price"], [("", ""), ("Coca Cola 1L", "2.5")])
    rows = parse_workbook(content)
    assert len(rows) == 1
    assert rows[0]["product_name"] == "Coca Cola 1L"


@pytest.mark.parametrize(
    ("raw", "expected"),
    [("1,99", "1.99"), ("1.99", "1.99"), ("1.234,56", "1234.56"), ("1234.56", "1234.56")],
)
def test_parse_decimal_handles_european_and_us_notation(raw, expected):
    from decimal import Decimal

    assert _parse_decimal(raw, field="price") == Decimal(expected)


def test_parse_decimal_rejects_negative_and_garbage():
    with pytest.raises(ValueError):
        _parse_decimal("-1", field="price")
    with pytest.raises(ValueError):
        _parse_decimal("abc", field="price")


def test_parse_decimal_rejects_values_that_would_overflow_the_price_column():
    # regular_price/promo_price are Numeric(10, 2); an out-of-range value must
    # fail cleanly here rather than reach Postgres as an unhandled overflow.
    with pytest.raises(ValueError):
        _parse_decimal("99999999999999999999999999.99", field="price")


def test_normalize_row_rejects_price_that_would_overflow_the_price_column():
    _, _, errors = _normalize_row(
        {"row": 2, "product_name": "Test", "price": "999999999999.99"}, market_currency="EUR"
    )
    assert any("price" in error for error in errors)


def test_normalize_row_requires_product_name():
    _, _, errors = _normalize_row({"row": 2}, market_currency="EUR")
    assert any("product_name" in error for error in errors)


def test_normalize_row_parses_legacy_package_strings():
    for value, expected_amount, expected_unit in (("100gr", "100", "g"), ("1lt", "1", "l"), ("33cl", "33", "cl")):
        _, normalized, errors = _normalize_row(
            {"row": 2, "product_name": "Test", "package_size": value}, market_currency="EUR"
        )
        assert errors == []
        assert normalized["package_amount"] == expected_amount
        assert normalized["package_unit"] == expected_unit


def test_normalize_row_rejects_unsupported_currency():
    _, _, errors = _normalize_row(
        {"row": 2, "product_name": "Test", "currency": "XYZ"}, market_currency="EUR"
    )
    assert any("para birimi" in error for error in errors)


def test_normalize_row_defaults_currency_to_market_currency():
    _, normalized, errors = _normalize_row({"row": 2, "product_name": "Test"}, market_currency="TRY")
    assert errors == []
    assert normalized["currency"] == "TRY"


def test_normalize_row_requires_amount_and_unit_together():
    _, _, errors = _normalize_row(
        {"row": 2, "product_name": "Test", "package_amount": "100"}, market_currency="EUR"
    )
    assert any("birlikte" in error for error in errors)


def test_normalize_row_rejects_non_http_image_url():
    _, normalized, errors = _normalize_row(
        {"row": 2, "product_name": "Test", "image_url": "javascript:alert(1)"}, market_currency="EUR"
    )
    assert any("image_url" in error for error in errors)
    assert normalized["image_url"] is None
